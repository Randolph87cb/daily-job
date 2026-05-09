from __future__ import annotations

import argparse
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

try:
    from openpyxl import load_workbook
except ModuleNotFoundError as exc:  # pragma: no cover - environment guard
    raise SystemExit(
        "缺少 openpyxl。先运行: python -m pip install openpyxl"
    ) from exc


SPECIAL_TWO_HOUR_STUDENTS = {"彭泓予", "聂楚恒"}
CLASS_ORDER = {"CSP-J": 0, "进阶班": 1, "一对一": 2}
WEEKDAY_MAP = {
    "星期一": 0,
    "星期二": 1,
    "星期三": 2,
    "星期四": 3,
    "星期五": 4,
    "星期六": 5,
    "星期日": 6,
    "星期天": 6,
}
WEEKDAY_CHARS = "一二三四五六日"
INACTIVE_STATUSES = {"", "退费", "退学", "休学", "结课", "换班/停"}


@dataclass
class Entry:
    teacher: str
    category: str
    class_name: str
    class_datetime: datetime
    attendee_count: int
    duration_hours: int
    remark: str | None
    slot_weekday: int
    time_key: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="从签到表生成课时统计表，并按项目规则统一排序和字段。"
    )
    parser.add_argument("--month", type=int, required=True, help="要提取的月份，例如 4")
    parser.add_argument(
        "--reference", type=Path, required=True, help="课时统计模板路径（.xlsx）"
    )
    parser.add_argument("--output", type=Path, required=True, help="输出文件路径（.xlsx）")
    parser.add_argument(
        "--sources",
        type=Path,
        nargs="+",
        required=True,
        help="一个或多个签到表路径",
    )
    return parser.parse_args()


def norm_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def norm_teacher(value: object) -> str:
    teacher = norm_text(value)
    if teacher in {"彬彬", "曹彬"}:
        return "曹彬"
    return teacher


def norm_class(value: object) -> str:
    text = norm_text(value)
    compact = (
        text.replace("—", "-")
        .replace("（", "(")
        .replace("）", ")")
        .replace(" ", "")
    )
    compact = re.sub(r"\s+", "", compact)
    if "CSP" in compact:
        return "CSP-J"
    if "进阶" in compact:
        return "进阶班"
    if "一对一" in compact:
        return "一对一"
    return text


def category_for_class(class_name: str) -> str:
    if class_name == "一对一":
        return "常规一对一课程"
    return "常规班课"


def parse_date_header(value: object) -> datetime | None:
    if isinstance(value, datetime):
        return value

    text = norm_text(value)
    if not text:
        return None

    match = re.search(
        r"(\d{4})[/-](\d{1,2})[/-](\d{1,2}).*?调\s*(\d{1,2})[/-](\d{1,2})",
        text,
    )
    if match:
        year, _src_month, _src_day, month, day = map(int, match.groups())
        return datetime(year, month, day)

    for pattern in (r"(\d{4})-(\d{1,2})-(\d{1,2})", r"(\d{4})/(\d{1,2})/(\d{1,2})"):
        match = re.search(pattern, text)
        if match:
            year, month, day = map(int, match.groups())
            return datetime(year, month, day)

    return None


def parse_time_range(text: str) -> tuple[int, int, int] | None:
    clean = norm_text(text).replace("：", ":").replace("（", "(").replace("）", ")")
    match = re.search(r"(\d{1,2}:\d{2})\s*-\s*(\d{1,2}:\d{2})", clean)
    if not match:
        return None

    start_text, end_text = match.groups()
    start_hour, start_minute = map(int, start_text.split(":"))
    end_hour, end_minute = map(int, end_text.split(":"))
    duration_minutes = end_hour * 60 + end_minute - start_hour * 60 - start_minute
    return start_hour, start_minute, duration_minutes // 60


def parse_attendance_sheet(path: Path, month: int) -> list[Entry]:
    workbook = load_workbook(path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    entries: list[Entry] = []
    current_dates: dict[int, datetime] = {}
    row = 1
    found_block = False

    while row <= worksheet.max_row:
        col_a = worksheet.cell(row, 1).value
        col_b = worksheet.cell(row, 2).value
        col_c = worksheet.cell(row, 3).value

        if col_a == "上课日期":
            current_dates = {}
            for col in range(8, worksheet.max_column + 1):
                parsed_date = parse_date_header(worksheet.cell(row, col).value)
                if parsed_date is not None:
                    current_dates[col] = parsed_date
            row += 1
            continue

        if col_b and col_c:
            lines = [piece.strip() for piece in str(col_b).splitlines() if piece.strip()]
            time_bits = parse_time_range(" ".join(lines))
            if time_bits:
                found_block = True
                start_hour, start_minute, duration_hours = time_bits
                class_name = norm_class(lines[0])
                teacher = norm_teacher(col_c)
                schedule_weekday = WEEKDAY_MAP.get(norm_text(col_a))
                students: list[tuple[str, dict[int, object]]] = []
                next_row = row

                while next_row <= worksheet.max_row:
                    if next_row != row and worksheet.cell(next_row, 1).value == "上课日期":
                        break
                    if (
                        next_row != row
                        and worksheet.cell(next_row, 2).value
                        and worksheet.cell(next_row, 3).value
                    ):
                        break

                    student_value = worksheet.cell(next_row, 4).value
                    if student_value:
                        student_name = norm_text(student_value).split("/")[0].strip()
                        status_map = {
                            col: worksheet.cell(next_row, col).value for col in current_dates
                        }
                        students.append((student_name, status_map))
                    next_row += 1

                for col, base_date in current_dates.items():
                    if base_date.month != month:
                        continue

                    attendee_count = 0
                    remarks: list[str] = []
                    for student_name, status_map in students:
                        status = norm_text(status_map.get(col))
                        if "✅" in status or status == "请假已补":
                            attendee_count += 1
                        if student_name in SPECIAL_TWO_HOUR_STUDENTS and "✅" in status:
                            remarks.append(f"{student_name}2小时")

                    if attendee_count == 0:
                        continue

                    class_datetime = base_date.replace(hour=start_hour, minute=start_minute)
                    slot_weekday = (
                        schedule_weekday if schedule_weekday is not None else class_datetime.weekday()
                    )
                    entries.append(
                        Entry(
                            teacher=teacher,
                            category=category_for_class(class_name),
                            class_name=class_name,
                            class_datetime=class_datetime,
                            attendee_count=attendee_count,
                            duration_hours=duration_hours,
                            remark="；".join(remarks) or None,
                            slot_weekday=slot_weekday,
                            time_key=f"{start_hour:02d}:{start_minute:02d}",
                        )
                    )

                row = next_row
                continue

        row += 1

    if not found_block:
        raise ValueError(f"{path} 不是可识别的签到表结构。")

    return entries


def write_output(
    entries: list[Entry],
    source_notes: list[str],
    reference_path: Path,
    output_path: Path,
    month: int,
) -> None:
    workbook = load_workbook(reference_path)
    worksheet = workbook[workbook.sheetnames[0]]
    worksheet["A1"] = f"{month}月课时统计"

    for row in range(3, worksheet.max_row + 1):
        for col in range(1, 9):
            worksheet.cell(row, col).value = None

    for index, entry in enumerate(entries, start=3):
        worksheet.cell(index, 1).value = index - 2
        worksheet.cell(index, 2).value = entry.teacher
        worksheet.cell(index, 3).value = entry.category
        worksheet.cell(index, 4).value = entry.class_name
        worksheet.cell(index, 5).value = entry.class_datetime
        worksheet.cell(index, 5).number_format = worksheet["E3"].number_format
        worksheet.cell(index, 6).value = entry.attendee_count
        worksheet.cell(index, 7).value = entry.duration_hours
        worksheet.cell(index, 8).value = entry.remark

    note_sheet = workbook[workbook.sheetnames[1]]
    for row in range(1, max(note_sheet.max_row, 10) + 1):
        note_sheet.cell(row, 1).value = None

    note_sheet["A1"] = "说明"
    note_sheet["A2"] = "班级名称统一为：CSP-J、进阶班、一对一。"
    note_sheet["A3"] = "讲师姓名统一规则：彬彬、曹彬均记为曹彬。"
    note_sheet["A4"] = "排序规则：班级 -> 原排课时段 -> 日期，同一时段连续展示。"
    note_sheet["A5"] = "备注规则：彭泓予、聂楚恒仅在该次课状态为✅、实际到课时标记 2 小时备注。"
    note_sheet["A6"] = "统计规则补充：请假已补计入当次学员数量，同时保留对应补课备注。"
    note_sheet["A7"] = source_notes[0] if len(source_notes) > 0 else ""
    note_sheet["A8"] = source_notes[1] if len(source_notes) > 1 else ""
    note_sheet["A9"] = "特殊日期处理：调课表头按实际日期统计，但仍归入原排课班次分组。"
    note_sheet["A10"] = f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"

    if len(workbook.sheetnames) >= 3:
        extra_sheet = workbook[workbook.sheetnames[2]]
        for row in extra_sheet.iter_rows():
            for cell in row:
                cell.value = None

    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        workbook.save(output_path)
    except PermissionError as exc:
        raise SystemExit(
            f"无法写入 {output_path}。文件可能正被 Excel 占用，先关闭文件或改用新的输出路径。"
        ) from exc


def main() -> int:
    args = parse_args()
    for path in [args.reference, *args.sources]:
        if not path.exists():
            raise SystemExit(f"文件不存在：{path}")

    entries: list[Entry] = []
    source_notes: list[str] = []
    for source in args.sources:
        parsed_entries = parse_attendance_sheet(source, args.month)
        entries.extend(parsed_entries)
        source_notes.append(
            f"{source.name}：按签到表结构解析，得到 {len(parsed_entries)} 条 {args.month} 月明细。"
        )

    entries.sort(
        key=lambda item: (
            CLASS_ORDER.get(item.class_name, 99),
            item.slot_weekday,
            item.time_key,
            item.class_datetime,
        )
    )

    write_output(entries, source_notes, args.reference, args.output, args.month)

    print(f"已生成：{args.output}")
    print(f"总记录数：{len(entries)}")
    for index, entry in enumerate(entries, start=1):
        weekday = WEEKDAY_CHARS[entry.slot_weekday]
        print(
            f"{index}. {entry.class_name} 周{weekday} {entry.time_key} "
            f"{entry.class_datetime:%Y-%m-%d %H:%M} {entry.attendee_count}人 "
            f"{entry.remark or ''}".rstrip()
        )

    return 0


if __name__ == "__main__":
    sys.exit(main())
